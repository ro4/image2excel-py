from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import sys

def rgb_to_hex(rgb):
    return '%02x%02x%02x' % rgb

def image_to_excel(image_path, output_path, progress_callback=None):
    # Open image file
    img = Image.open(image_path)
    
    # Get image dimensions
    width, height = img.size
    total_pixels = width * height
    processed_pixels = 0
    
    # Create new Excel workbook
    wb = Workbook()
    ws = wb.active
    
    # Set column width and row height to make cells square-like (1 char width ≈8px, 8pt height ≈12px)
    for x in range(width):
        ws.column_dimensions[ws.cell(row=1, column=x+1).column_letter].width = 1.0  # 1.0字符≈7像素（1字符=7像素）
        # 列宽7像素（1.0字符） / 行高6点≈8像素（1点=1.333像素） 实现0.875:1近似正方形
    
    for y in range(height):
        ws.row_dimensions[y+1].height = 6  # 6点≈8像素（1点=1.333像素）
    
    # Iterate through image pixels
    for y in range(height):
        # 行高已统一设置为6点（外层循环）
        
        for x in range(width):
            # Get pixel RGB value
            pixel = img.getpixel((x, y))
            # 处理不同的图像模式
            if img.mode == 'L':  # 灰度图
                pixel = (pixel, pixel, pixel)
            elif img.mode == 'LA':  # 灰度图 + Alpha
                gray, alpha = pixel
                alpha = alpha/255
                if alpha == 0:
                    pixel = (255, 255, 255)
                else:
                    gray_with_alpha = int(gray*alpha + 255*(1-alpha))
                    pixel = (gray_with_alpha, gray_with_alpha, gray_with_alpha)
            elif img.mode == 'RGBA':
                r, g, b, alpha = pixel
                alpha = alpha/255
                if alpha == 0:
                    pixel = (255, 255, 255)
                else:
                    pixel = (
                        int(r*alpha + 255*(1-alpha)),
                        int(g*alpha + 255*(1-alpha)),
                        int(b*alpha + 255*(1-alpha))
                    )
            elif isinstance(pixel, int):  # 其他单通道图像
                pixel = (pixel, pixel, pixel)
            elif len(pixel) == 2:  # 其他双通道图像
                v1, v2 = pixel
                pixel = (v1, v1, v1)
            
            # Convert RGB to hex color code
            hex_color = rgb_to_hex(pixel)
            
            # Set cell background color
            cell = ws.cell(row=y+1, column=x + 1)
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
            
            # Update progress
            processed_pixels += 1
            if progress_callback and processed_pixels % 100 == 0:  # 每处理100个像素更新一次进度
                progress = (processed_pixels / total_pixels) * 100
                progress_callback(progress)
    
    # 保存Excel文件
    wb.save(output_path)
    
    # 确保进度显示100%
    if progress_callback:
        progress_callback(100)

def main():
    if len(sys.argv) != 3:
        print("Usage: python image2excel.py <input_image> <output_excel>")
        print("用法：python image2excel.py <输入图片> <输出Excel文件>")
        sys.exit(1)
    
    image_path = sys.argv[1]
    output_path = sys.argv[2]
    
    try:
        image_to_excel(image_path, output_path)
        print(f"Successfully converted {image_path} to {output_path}")
        print(f"成功将 {image_path} 转换为 {output_path}")
    except Exception as e:
        print(f"Error: {str(e)}")
        print(f"错误：{str(e)}")
        sys.exit(1)

if __name__ == '__main__':
    main()